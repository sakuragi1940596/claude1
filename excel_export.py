import os
from io import BytesIO
from openpyxl import load_workbook

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '許可申請書.xlsx')
OFFICERS_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '役員の一覧（法人用）.xlsx')
OFFICES_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '営業所一覧（新規）.xlsx')
TECHNICIANS_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '営業所技術者等一覧.xlsx')

# 建設業29業種コード
BUSINESS_TYPES = [
    ('土', 'civil_engineering'),
    ('建', 'architecture'),
    ('大', 'carpentry'),
    ('左', 'plastering'),
    ('と', 'masonry'),
    ('石', 'stone'),
    ('屋', 'roofing'),
    ('電', 'electrical'),
    ('管', 'plumbing'),
    ('タ', 'tiling'),
    ('鋼', 'steel'),
    ('筋', 'rebar'),
    ('舗', 'paving'),
    ('しゆ', 'dredging'),
    ('板', 'sheet_metal'),
    ('ガ', 'glass'),
    ('塗', 'painting'),
    ('防', 'waterproofing'),
    ('内', 'interior'),
    ('機', 'machinery'),
    ('絶', 'insulation'),
    ('通', 'telecom'),
    ('園', 'landscaping'),
    ('井', 'well'),
    ('具', 'fixtures'),
    ('水', 'water'),
    ('消', 'fire_protection'),
    ('清', 'cleaning'),
    ('解', 'demolition'),
]

# 業種に対応するExcelセル列（結合セルの開始列）
BUSINESS_TYPE_COLUMNS = [
    'AR', 'AV', 'AZ', 'BD', 'BH', 'BL', 'BP', 'BT', 'BX', 'CB',
    'CF', 'CJ', 'CN', 'CR', 'CV', 'CZ', 'DD', 'DH', 'DL', 'DP',
    'DT', 'DX', 'EB', 'EF', 'EJ', 'EN', 'ER', 'EV', 'EZ',
]


def _fill_cells(ws, cells, text):
    """文字列を1文字ずつ指定セルに書き込む"""
    if not text:
        return
    for i, char in enumerate(text):
        if i >= len(cells):
            break
        ws[cells[i]] = char


def _fill_digits(ws, cells, number_str):
    """数字文字列を右詰めで1桁ずつセルに書き込む"""
    if not number_str:
        return
    digits = str(number_str).strip()
    # 右詰め
    start = len(cells) - len(digits)
    for i, d in enumerate(digits):
        pos = start + i
        if 0 <= pos < len(cells):
            ws[cells[pos]] = d


def generate_excel(application, customer):
    """申請データからExcelファイルを生成してバイトデータを返す"""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb['様式第一号']

    # ============================================================
    # 申請日（行9: 令和__年__月__日）
    # 年: EZ9（結合セル1つ）, 月: FJ9, 日: FU9
    # ============================================================
    if application.get('application_date'):
        parts = application['application_date'].split('-')
        if len(parts) == 3:
            year = int(parts[0]) - 2018  # 西暦→令和
            ws['EZ9'] = str(year)
            ws['FJ9'] = str(int(parts[1]))
            ws['FU9'] = str(int(parts[2]))

    # ============================================================
    # 項番01:入力不要
    # ============================================================

    # ============================================================
    # 項番02:入力不要
    # ============================================================
    # 許可の有効期間の調整（1:する 2:しない）
    if application.get('validity_adjustment'):
        ws['FE25'] = str(application['validity_adjustment'])
    # ============================================================
    # 項番03:入力不要
    # ============================================================
  
    # ============================================================
    # 項番04: 許可を受けようとする建設業（行34）
    # ============================================================
    if application.get('business_types'):
        selected = application['business_types'].split(',')
        for i, (label, code) in enumerate(BUSINESS_TYPES):
            if code in selected:
                col = BUSINESS_TYPE_COLUMNS[i]
                ws[f'{col}34'] = '1'

    # ============================================================
    # 項番05: 既に許可を受けている建設業（行37）
    # ============================================================
    if application.get('existing_business_types'):
        selected = application['existing_business_types'].split(',')
        for i, (label, code) in enumerate(BUSINESS_TYPES):
            if code in selected:
                col = BUSINESS_TYPE_COLUMNS[i]
                ws[f'{col}37'] = '1'

    # ============================================================
    # 項番06: 商号又は名称のフリガナ（行40, 20文字分）
    # ============================================================
    name_kana_cells = [
        'AR40', 'AY40', 'BF40', 'BM40', 'BT40', 'CA40', 'CH40', 'CO40', 'CV40', 'DC40',
        'DJ40', 'DQ40', 'DX40', 'EE40', 'EL40', 'ES40', 'EZ40', 'FG40', 'FN40', 'FU40',
    ]
    name_kana = customer.get('name_kana') or ''
    if customer.get('corporation_type') and int(customer['corporation_type']) == 1:
        # 法人の場合、法人格のフリガナを除去
        corp_kana_list = [
            'カブシキガイシャ', 'ユウゲンガイシャ', 'ゴウドウガイシャ', 'ゴウシガイシャ',
            'ゴウメイガイシャ', 'イッパンシャダンホウジン', 'イッパンザイダンホウジン',
            'コウエキシャダンホウジン', 'コウエキザイダンホウジン',
            'トクテイヒエイリカツドウホウジン', 'シャカイフクシホウジン',
            'ガッコウホウジン', 'イリョウホウジン',
        ]
        for corp_kana in corp_kana_list:
            name_kana = name_kana.replace(corp_kana, '').strip()
    _fill_cells(ws, name_kana_cells, name_kana)

    # ============================================================
    # 項番07: 商号又は名称（行46, 20文字分）
    # ============================================================
    name_cells = [
        'AR46', 'AY46', 'BF46', 'BM46', 'BT46', 'CA46', 'CH46', 'CO46', 'CV46', 'DC46',
        'DJ46', 'DQ46', 'DX46', 'EE46', 'EL46', 'ES46', 'EZ46', 'FG46', 'FN46', 'FU46',
    ]
    name = customer.get('name') or ''
    if customer.get('corporation_type') and int(customer['corporation_type']) == 1:
        # 法人格を略称に置換
        corp_replacements = [
            ('株式会社', '（株）'),
            ('有限会社', '（有）'),
            ('合同会社', '（合）'),
            ('合資会社', '（資）'),
            ('合名会社', '（名）'),
            ('一般社団法人', '（一社）'),
            ('一般財団法人', '（一財）'),
            ('公益社団法人', '（公社）'),
            ('公益財団法人', '（公財）'),
            ('特定非営利活動法人', '（特非）'),
            ('社会福祉法人', '（福）'),
            ('学校法人', '（学）'),
            ('医療法人', '（医）'),
        ]
        for full, short in corp_replacements:
            name = name.replace(full, short)
    _fill_cells(ws, name_cells, name)

    # ============================================================
    # 項番08: 代表者又は個人の氏名のフリガナ（行52, 20文字分）
    # ============================================================
    rep_kana_cells = [
        'AR52', 'AY52', 'BF52', 'BM52', 'BT52', 'CA52', 'CH52', 'CO52', 'CV52', 'DC52',
        'DJ52', 'DQ52', 'DX52', 'EE52', 'EL52', 'ES52', 'EZ52', 'FG52', 'FN52', 'FU52',
    ]
    _fill_cells(ws, rep_kana_cells, customer.get('representative_kana'))

    # ============================================================
    # 項番09: 代表者又は個人の氏名（行55, 10文字分）
    # ============================================================
    rep_name_cells = [
        'AR55', 'AY55', 'BF55', 'BM55', 'BT55', 'CA55', 'CH55', 'CO55', 'CV55', 'DC55',
    ]
    _fill_cells(ws, rep_name_cells, customer.get('representative'))

    # ============================================================
    # 項番10: 市区町村コード（行58, 5桁: AR58,AV58,AZ58,BD58,BH58）
    # ============================================================
    city_code_cells = ['AR58', 'AV58', 'AZ58', 'BD58', 'BH58']
    if application.get('city_code'):
        _fill_digits(ws, city_code_cells, application['city_code'])

    # 都道府県名（CF58 結合セル）
    if customer.get('prefecture'):
        ws['CF58'] = customer['prefecture']

    # 市区町村名（EK58 結合セル）
    if customer.get('city'):
        ws['EK58'] = customer['city']

    # ============================================================
    # 項番11: 主たる営業所の所在地（行61, 20文字分）
    # ============================================================
    address_cells = [
        'AR61', 'AY61', 'BF61', 'BM61', 'BT61', 'CA61', 'CH61', 'CO61', 'CV61', 'DC61',
        'DJ61', 'DQ61', 'DX61', 'EE61', 'EL61', 'ES61', 'EZ61', 'FG61', 'FN61', 'FU61',
    ]
    address = customer.get('address', '') or ''
    address = address.replace('丁目', '-').replace('番地', '-').replace('番', '-').replace('号', '')
    # 末尾のハイフンを除去
    address = address.rstrip('-')
    _fill_cells(ws, address_cells, address)

    # ============================================================
    # 項番12: 郵便番号（3桁: AR67,AV67,AZ67 / 4桁: BH67,BL67,BP67,BT67）
    # ============================================================
    if customer.get('postal_code'):
        parts = customer['postal_code'].replace('ー', '-').replace('－', '-').split('-')
        if len(parts) == 2:
            _fill_cells(ws, ['AR67', 'AV67', 'AZ67'], parts[0])
            _fill_cells(ws, ['BH67', 'BL67', 'BP67', 'BT67'], parts[1])

    # 電話番号（DD67〜EZ67, 最大12文字）
    phone_cells = [
        'DD67', 'DH67', 'DL67', 'DP67', 'DT67', 'DX67',
        'EB67', 'EF67', 'EJ67', 'EN67', 'ER67', 'EV67', 'EZ67',
    ]
    if customer.get('phone'):
        _fill_cells(ws, phone_cells, customer['phone'].replace('-', '').replace('ー', ''))

    # FAX番号（BQ70 結合セル）
    if customer.get('fax'):
        ws['BQ70'] = customer['fax']
    else:
        ws['BQ70'] = 'なし'

    # ============================================================
    # 項番13: 法人又は個人の別
    # ============================================================
    if customer.get('corporation_type'):
        ws['AR75'] = str(customer['corporation_type'])

    # 資本金額（千円）（行75: BT75,BX75,CB75,CF75,CJ75,CN75,CR75,CV75,CZ75）
    capital_cells = ['BT75', 'BX75', 'CB75', 'CF75', 'CJ75', 'CN75', 'CR75', 'CV75', 'CZ75']
    if customer.get('capital_amount'):
        _fill_digits(ws, capital_cells, customer['capital_amount'])

    # 法人番号（13桁: DX75,EB75,EF75,EJ75,EN75,ER75,EV75,EZ75,FD75,FH75,FL75,FP75,FT75）
    corp_number_cells = [
        'DX75', 'EB75', 'EF75', 'EJ75', 'EN75', 'ER75', 'EV75',
        'EZ75', 'FD75', 'FH75', 'FL75', 'FP75', 'FT75',
    ]
    if customer.get('corporate_number'):
        _fill_cells(ws, corp_number_cells, customer['corporate_number'])

    # ============================================================
    # 項番14: 兼業の有無
    # ============================================================
    if application.get('side_business'):
        ws['AR78'] = str(application['side_business'])
        if int(application['side_business']) == 2:
            ws['CG79'] = 'なし'
        elif application.get('side_business_type'):
            ws['CG79'] = application['side_business_type']

    # ============================================================
    # 項番15: 許可換えの区分
    # ============================================================
    if application.get('permit_transfer_category'):
        ws['AL82'] = str(application['permit_transfer_category'])

    # ============================================================
    # 項番16: 旧許可番号
    # ============================================================
    # 旧許可番号（6桁: CZ88,DD88,DH88,DL88,DP88,DT88）
    old_permit_cells = ['CZ88', 'DD88', 'DH88', 'DL88', 'DP88', 'DT88']
    if application.get('old_permit_number'):
        _fill_digits(ws, old_permit_cells, application['old_permit_number'])
    if application.get('old_permit_year'):
        _fill_digits(ws, ['EL88', 'EP88'], application['old_permit_year'])
    if application.get('old_permit_month'):
        _fill_digits(ws, ['EW88', 'FA88'], application['old_permit_month'])
    if application.get('old_permit_day'):
        _fill_digits(ws, ['FH88', 'FL88'], application['old_permit_day'])

    # ============================================================
    # 申請者・代理人（行11-16）
    # ============================================================
    # 申請者住所（所在地）
    if application.get('applicant_address'):
        ws['DK11'] = application['applicant_address']
    # 申請者氏名（法人名）
    if application.get('applicant_name'):
        ws['DK12'] = application['applicant_name']
    # 代表者氏名（法人の場合：肩書＋氏名）
    if customer.get('corporation_type') and int(customer['corporation_type']) == 1:
        if customer.get('representative'):
            title = customer.get('representative_title', '')
            name = customer['representative']
            ws['DK13'] = f"{title}　{name}" if title else name
    # 申請者代理人
    if application.get('proxy_name'):
        ws['DL16'] = application['proxy_name']

    # ============================================================
    # 連絡先（行97-105）
    # ============================================================
    if application.get('contact_organization'):
        ws['B97'] = application['contact_organization']
    if application.get('contact_name'):
        ws['AZ97'] = application['contact_name']
    if application.get('contact_phone'):
        ws['B101'] = application['contact_phone']
    if application.get('contact_fax'):
        ws['B105'] = application['contact_fax']

    # バイトデータとして返す
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# 営業所一覧用の業種列（許可申請書とは異なるオフセット）
OFFICES_BIZ_TYPE_COLUMNS = [
    'AS', 'AW', 'BA', 'BE', 'BI', 'BM', 'BQ', 'BU', 'BY', 'CC',
    'CG', 'CK', 'CO', 'CS', 'CW', 'DA', 'DE', 'DI', 'DM', 'DQ',
    'DU', 'DY', 'EC', 'EG', 'EK', 'EO', 'ES', 'EW', 'FA',
]

# 従たる営業所の名前セル（20文字：各6列幅の結合セル）
_BRANCH_NAME_CELL_COLS = [
    'AS', 'AZ', 'BG', 'BN', 'BU', 'CB', 'CI', 'CP', 'CW', 'DD',
    'DK', 'DR', 'DY', 'EF', 'EM', 'ET', 'FA', 'FH', 'FO', 'FV',
]


def generate_offices_list_excel(application, customer, offices):
    """営業所一覧表（新規）のExcelを生成してバイトデータを返す"""
    wb = load_workbook(OFFICES_TEMPLATE_PATH)
    ws = wb['様式第一号別紙二（１）']

    # ============================================================
    # 項番82: 許可番号（行13-14）
    # ============================================================
    # 許可番号（6桁: DA13〜DU13）
    permit_cells = ['DA13', 'DE13', 'DI13', 'DM13', 'DQ13', 'DU13']
    if application.get('permit_number'):
        _fill_digits(ws, permit_cells, application['permit_number'])

    # 許可年月日
    if application.get('permit_year'):
        _fill_digits(ws, ['EM13', 'EQ13'], application['permit_year'])
    if application.get('permit_month'):
        _fill_digits(ws, ['EX13', 'FB13'], application['permit_month'])
    if application.get('permit_day'):
        _fill_digits(ws, ['FI13', 'FM13'], application['permit_day'])

    # ============================================================
    # 主たる営業所（行18-29）
    # ============================================================
    main_office = None
    for o in offices:
        if o.get('office_type') == 1:
            main_office = o
            break

    # フリガナ（BE19 結合セル）
    main_name_kana = (main_office.get('name_kana', '') if main_office else '') or 'ホンテン'
    ws['BE19'] = main_name_kana

    # 名称（BE21 結合セル）
    main_name = (main_office.get('name', '') if main_office else '') or '本店'
    ws['BE21'] = main_name

    if main_office and main_office.get('business_types'):
        selected = main_office['business_types'].split(',')
        for i, (label, code) in enumerate(BUSINESS_TYPES):
            if code in selected:
                col = OFFICES_BIZ_TYPE_COLUMNS[i]
                ws[f'{col}25'] = '1'

    # ============================================================
    # 従たる営業所（最大2件）
    # ============================================================
    branch_offices = [o for o in offices if o.get('office_type') == 2]

    branch_configs = [
        {   # 従たる営業所1
            'kana_cell': 'BD33', 'name_row1': 36, 'name_row2': 39,
            'city_code_cells': ['AS43', 'AW43', 'BA43', 'BE43', 'BI43'],
            'prefecture_cell': 'CH43', 'city_cell': 'EO43',
            'addr_row1': 46, 'addr_row2': 49,
            'postal_first': ['AS52', 'AW52', 'BA52'],
            'postal_last': ['BI52', 'BM52', 'BQ52', 'BU52'],
            'phone_cells': [
                'DE52', 'DI52', 'DM52', 'DQ52', 'DU52', 'DY52',
                'EC52', 'EG52', 'EK52', 'EO52', 'ES52', 'EW52', 'FA52',
            ],
            'biz_row': 56,
        },
        {   # 従たる営業所2
            'kana_cell': 'BD64', 'name_row1': 67, 'name_row2': 70,
            'city_code_cells': ['AS74', 'AW74', 'BA74', 'BE74', 'BI74'],
            'prefecture_cell': 'CH74', 'city_cell': 'EO74',
            'addr_row1': 77, 'addr_row2': 80,
            'postal_first': ['AS83', 'AW83', 'BA83'],
            'postal_last': ['BI83', 'BM83', 'BQ83', 'BU83'],
            'phone_cells': [
                'DE83', 'DI83', 'DM83', 'DQ83', 'DU83', 'DY83',
                'EC83', 'EG83', 'EK83', 'EO83', 'ES83', 'EW83', 'FA83',
            ],
            'biz_row': 87,
        },
    ]

    for idx, cfg in enumerate(branch_configs):
        if idx >= len(branch_offices):
            break
        office = branch_offices[idx]

        # フリガナ（結合セル）
        if office.get('name_kana'):
            ws[cfg['kana_cell']] = office['name_kana']

        # 名称（1行目20文字 + 2行目20文字 = 最大40文字）
        name = office.get('name', '') or ''
        name_cells_1 = [f'{col}{cfg["name_row1"]}' for col in _BRANCH_NAME_CELL_COLS]
        name_cells_2 = [f'{col}{cfg["name_row2"]}' for col in _BRANCH_NAME_CELL_COLS]
        _fill_cells(ws, name_cells_1, name[:20])
        if len(name) > 20:
            _fill_cells(ws, name_cells_2, name[20:40])

        # 市区町村コード（5桁）
        if office.get('city_code'):
            _fill_digits(ws, cfg['city_code_cells'], office['city_code'])

        # 都道府県名
        if office.get('prefecture'):
            ws[cfg['prefecture_cell']] = office['prefecture']

        # 市区町村名
        if office.get('city'):
            ws[cfg['city_cell']] = office['city']

        # 所在地（1行目20文字 + 2行目20文字 = 最大40文字）
        addr_cells_1 = [f'{col}{cfg["addr_row1"]}' for col in _BRANCH_NAME_CELL_COLS]
        addr_cells_2 = [f'{col}{cfg["addr_row2"]}' for col in _BRANCH_NAME_CELL_COLS]
        address = office.get('address', '') or ''
        address = address.replace('丁目', '-').replace('番地', '-').replace('番', '-').replace('号', '')
        address = address.rstrip('-')
        _fill_cells(ws, addr_cells_1, address[:20])
        if len(address) > 20:
            _fill_cells(ws, addr_cells_2, address[20:40])

        # 郵便番号
        if office.get('postal_code'):
            parts = office['postal_code'].replace('ー', '-').replace('－', '-').split('-')
            if len(parts) == 2:
                _fill_cells(ws, cfg['postal_first'], parts[0])
                _fill_cells(ws, cfg['postal_last'], parts[1])

        # 電話番号
        if office.get('phone'):
            phone = office['phone'].replace('-', '').replace('ー', '')
            _fill_cells(ws, cfg['phone_cells'], phone)

        # 営業しようとする建設業（29業種）
        if office.get('business_types'):
            selected = office['business_types'].split(',')
            for i, (label, code) in enumerate(BUSINESS_TYPES):
                if code in selected:
                    col = OFFICES_BIZ_TYPE_COLUMNS[i]
                    ws[f'{col}{cfg["biz_row"]}'] = '1'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_technicians_excel(technicians, application_date=None):
    """営業所技術者等一覧表（様式第一号別紙四）のExcelを生成してバイトデータを返す"""
    from openpyxl.styles import Alignment, Font, Border, Side

    wb = load_workbook(TECHNICIANS_TEMPLATE_PATH)
    ws = wb['様式第一号別紙四']

    # 日付（行8: 結合セル B8:EW8）
    if application_date:
        parts = application_date.split('-')
        if len(parts) == 3:
            year = int(parts[0]) - 2018  # 西暦→令和
            month = int(parts[1])
            day = int(parts[2])
            ws['B8'] = f'令和　{year}　年　{month}　月　{day}　日'

    # データエリアの大きな結合セルを解除
    merges_to_remove = []
    for merged in ws.merged_cells.ranges:
        if merged.min_row == 13 and merged.max_row == 81:
            merges_to_remove.append(str(merged))
    for m in merges_to_remove:
        ws.unmerge_cells(m)

    # スタイル定義
    thin_border_side = Side(style='thin', color='000000')
    cell_font = Font(name='ＭＳ 明朝', size=10)
    cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 列範囲定義（データ4列）
    col_ranges = [
        (2, 37),    # B:AK  - 営業所の名称
        (38, 85),   # AL:CG - 営業所技術者等の氏名
        (86, 119),  # CH:DO - 建設工事の種類
        (120, 153), # DP:EW - 有資格区分
    ]

    # 技術者1名あたりの行数（3行分を使用）
    rows_per_tech = 3
    max_technicians = (81 - 13 + 1) // rows_per_tech  # 最大23名

    for i, tech in enumerate(technicians[:max_technicians]):
        start_row = 13 + i * rows_per_tech
        end_row = start_row + rows_per_tech - 1

        # 各列グループのセルを結合
        from openpyxl.utils import get_column_letter
        for col_start, col_end in col_ranges:
            merge_range = f'{get_column_letter(col_start)}{start_row}:{get_column_letter(col_end)}{end_row}'
            ws.merge_cells(merge_range)

        # 結合セルの先頭セルにスタイルを適用
        for col_idx, (col_start, col_end) in enumerate(col_ranges):
            cell = ws.cell(row=start_row, column=col_start)
            cell.font = cell_font
            cell.alignment = cell_alignment

        # 罫線を各行・列に適用（結合セル全体に罫線が表示されるように）
        for row in range(start_row, end_row + 1):
            for col_start, col_end in col_ranges:
                for col in range(col_start, col_end + 1):
                    cell = ws.cell(row=row, column=col)
                    border_left = thin_border_side if col == col_start else None
                    border_right = thin_border_side if col == col_end else None
                    border_top = thin_border_side if row == start_row else None
                    border_bottom = thin_border_side if row == end_row else None
                    cell.border = Border(
                        left=border_left or Side(),
                        right=border_right or Side(),
                        top=border_top or Side(),
                        bottom=border_bottom or Side(),
                    )

        # 行の高さを設定
        for row in range(start_row, end_row + 1):
            ws.row_dimensions[row].height = 18

        # ① 営業所の名称（B列）
        office_name = tech.get('office_name', '') or ''
        ws.cell(row=start_row, column=2).value = office_name

        # ② 営業所技術者等の氏名（AL列）- フリガナ＋改行＋氏名
        name = tech.get('name', '') or ''
        name_kana = tech.get('name_kana', '') or ''
        if name_kana:
            ws.cell(row=start_row, column=38).value = f'{name_kana}\n{name}'
        else:
            ws.cell(row=start_row, column=38).value = name

        # ③ 建設工事の種類（CH列）
        construction_types = tech.get('construction_types', '') or ''
        ws.cell(row=start_row, column=86).value = construction_types

        # ④ 有資格区分（DP列）- 建設工事の種類と対応するように改行
        qualifications = tech.get('qualifications', '') or ''
        ws.cell(row=start_row, column=120).value = qualifications

    # 未使用行の残りエリアを結合（技術者データの後ろ）
    used_rows = len(technicians[:max_technicians]) * rows_per_tech
    remaining_start = 13 + used_rows
    if remaining_start <= 81:
        from openpyxl.utils import get_column_letter
        for col_start, col_end in col_ranges:
            merge_range = f'{get_column_letter(col_start)}{remaining_start}:{get_column_letter(col_end)}81'
            ws.merge_cells(merge_range)
        # 残りエリアに罫線を適用
        for col_start, col_end in col_ranges:
            for row in range(remaining_start, 82):
                for col in range(col_start, col_end + 1):
                    cell = ws.cell(row=row, column=col)
                    border_left = thin_border_side if col == col_start else None
                    border_right = thin_border_side if col == col_end else None
                    border_top = thin_border_side if row == remaining_start else None
                    border_bottom = thin_border_side if row == 81 else None
                    cell.border = Border(
                        left=border_left or Side(),
                        right=border_right or Side(),
                        top=border_top or Side(),
                        bottom=border_bottom or Side(),
                    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_officers_excel(officers, application_date=None):
    """役員一覧表（様式第一号別紙一）のExcelを生成してバイトデータを返す"""
    from openpyxl.styles import Alignment, Font
    wb = load_workbook(OFFICERS_TEMPLATE_PATH)
    ws = wb['様式第一号別紙一']

    # ヘッダーのフリガナ復元（openpyxlでルビが消えるため改行テキストで再現）
    ws['B8'] = 'フリ\n氏'
    ws['AL8'] = 'ガナ\n名'
    for cell_ref in ['B8', 'AL8']:
        ws[cell_ref].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[cell_ref].font = Font(name='ＭＳ 明朝', size=12)

    # 日付（行5: FD5 結合セル）
    if application_date:
        parts = application_date.split('-')
        if len(parts) == 3:
            year = int(parts[0]) - 2018  # 西暦→令和
            month = int(parts[1])
            day = int(parts[2])
            ws['FD5'] = f'令和　{year}　年　{month}　月　{day}　日'

    # 役員データ（行9〜31、最大23名）
    for i, officer in enumerate(officers[:23]):
        row = 9 + i
        # フリガナ\n氏名 の2行構成
        last_kana = officer.get('last_name_kana', '') or ''
        first_kana = officer.get('first_name_kana', '') or ''
        last_name = officer.get('last_name', '') or ''
        first_name = officer.get('first_name', '') or ''
        ws[f'B{row}'] = f'{last_kana}\n{last_name}' if last_kana else last_name
        ws[f'AL{row}'] = f'{first_kana}\n{first_name}' if first_kana else first_name
        ws[f'BT{row}'] = officer.get('role', '')
        ws[f'EL{row}'] = officer.get('full_or_part', '')
        for cell_ref in [f'B{row}', f'AL{row}']:
            ws[cell_ref].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
